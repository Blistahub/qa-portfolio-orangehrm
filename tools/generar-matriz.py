# -*- coding: utf-8 -*-
"""Genera la matriz de casos de prueba en los tres formatos entregables.

    python tools/generar-matriz.py

Salida (en la raíz del repositorio):
    02-casos-de-prueba.md    → lectura directa en GitHub
    02-casos-de-prueba.csv   → intercambio y control de versiones legible
    02-casos-de-prueba.xlsx  → formato habitual de entrega al cliente

El objetivo de generar los tres desde `tools/casos.py` es evitar que diverjan:
en un proyecto real, una matriz en Excel y su copia en Markdown se
desincronizan al segundo cambio.
"""

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from casos import CASOS, COLUMNAS  # noqa: E402

RAIZ = Path(__file__).resolve().parent.parent
BASE = "02-casos-de-prueba"

ORDEN_PRIORIDAD = {"Alta": 0, "Media": 1, "Baja": 2}


def normaliza(valor):
    """«Tabla de decisión (R3)» → «Tabla de decisión», para agrupar en los resúmenes."""
    return valor.split(" (")[0]


def filas():
    """Devuelve los casos como listas de texto plano, con los pasos numerados."""
    for c in CASOS:
        cid, req, titulo, precond, pasos, datos, esperado, prio, tipo, tec, estado, bug = c
        pasos_txt = "\n".join(f"{i}. {p}" for i, p in enumerate(pasos, 1))
        yield [cid, req, titulo, precond, pasos_txt, datos, esperado,
               prio, tipo, tec, estado, bug or "—"]


# --------------------------------------------------------------------------- CSV
def generar_csv(destino):
    with open(destino, "w", newline="", encoding="utf-8-sig") as f:
        # lineterminator explícito: si no, csv escribe CRLF y el fichero deja de
        # ser idéntico entre Windows y Linux, rompiendo la comprobación de CI.
        w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL, lineterminator="\n")
        w.writerow(COLUMNAS)
        w.writerows(filas())
    print(f"  ✓ {destino.name}")


# -------------------------------------------------------------------------- XLSX
def generar_xlsx(destino):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    AZUL = "1F3A5F"
    GRIS = "F2F4F7"
    VERDE = "D6EFD8"
    ROJO = "FBD9D9"
    AMBAR = "FDF0CE"

    wb = Workbook()
    ws = wb.active
    ws.title = "Casos de prueba"

    borde = Border(*(Side(style="thin", color="D0D5DD"),) * 4)

    ws.append(COLUMNAS)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        celda.border = borde
    ws.row_dimensions[1].height = 32

    relleno_estado = {"Pasa": VERDE, "Falla": ROJO, "Bloqueado": AMBAR}

    for fila in filas():
        ws.append(fila)
        n = ws.max_row
        for celda in ws[n]:
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = borde
        if n % 2 == 0:
            for celda in ws[n]:
                celda.fill = PatternFill("solid", fgColor=GRIS)
        estado = fila[10]
        c_estado = ws.cell(row=n, column=11)
        c_estado.fill = PatternFill("solid", fgColor=relleno_estado[estado])
        c_estado.font = Font(bold=True)
        c_estado.alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=n, column=1).font = Font(bold=True)
        ws.row_dimensions[n].height = 96

    anchos = [10, 18, 40, 30, 46, 30, 46, 10, 18, 22, 12, 11]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ref = f"A1:{get_column_letter(len(COLUMNAS))}{ws.max_row}"
    tabla = Table(displayName="MatrizCasosPIM", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
    ws.add_table(tabla)
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False

    # --- Hoja de resumen -----------------------------------------------------
    rs = wb.create_sheet("Resumen")
    rs.sheet_view.showGridLines = False
    rs["A1"] = "Resumen de ejecución — módulo PIM"
    rs["A1"].font = Font(bold=True, size=14, color=AZUL)

    def bloque(titulo, datos, fila_inicio):
        rs.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True, color="FFFFFF")
        rs.cell(row=fila_inicio, column=1).fill = PatternFill("solid", fgColor=AZUL)
        rs.cell(row=fila_inicio, column=2, value="Casos").font = Font(bold=True, color="FFFFFF")
        rs.cell(row=fila_inicio, column=2).fill = PatternFill("solid", fgColor=AZUL)
        for i, (k, v) in enumerate(datos.items(), 1):
            rs.cell(row=fila_inicio + i, column=1, value=k)
            rs.cell(row=fila_inicio + i, column=2, value=v).alignment = Alignment(horizontal="center")
        return fila_inicio + len(datos) + 2

    def contar(indice, orden=None):
        d = {}
        for c in CASOS:
            clave = normaliza(c[indice])
            d[clave] = d.get(clave, 0) + 1
        if orden:
            return {k: d[k] for k in orden if k in d}
        return dict(sorted(d.items(), key=lambda x: -x[1]))

    f = 3
    f = bloque("Estado", contar(10, ["Pasa", "Falla", "Bloqueado"]), f)
    f = bloque("Prioridad", contar(7, ["Alta", "Media", "Baja"]), f)
    f = bloque("Tipo de prueba", contar(8), f)
    f = bloque("Técnica de diseño", contar(9), f)
    rs.column_dimensions["A"].width = 34
    rs.column_dimensions["B"].width = 12

    wb.save(destino)
    print(f"  ✓ {destino.name}")


# ---------------------------------------------------------------------------- MD
ICONO = {"Pasa": "🟢 Pasa", "Falla": "🔴 Falla", "Bloqueado": "🟡 Bloqueado"}

CABECERA_MD = """# 02 — Matriz de casos de prueba · Módulo PIM

**{total} casos** diseñados sobre el catálogo de [requisitos](00-requisitos.md), siguiendo la
estrategia del [plan de pruebas](01-plan-de-pruebas.md).

> Fichero generado automáticamente por `tools/generar-matriz.py` a partir de `tools/casos.py`.
> No editar a mano: los tres formatos entregables —`.md`, `.csv` y `.xlsx`— salen de la misma
> fuente para que no puedan divergir.

**Otros formatos:** [`02-casos-de-prueba.xlsx`](02-casos-de-prueba.xlsx) ·
[`02-casos-de-prueba.csv`](02-casos-de-prueba.csv)

| Estado | Casos | | Prioridad | Casos | | Técnica de diseño | Casos |
| ------ | ----: | - | --------- | ----: | - | ----------------- | ----: |
{resumen}

---

## Resumen de la matriz

| ID | Requisito | Título | Prioridad | Tipo | Técnica | Estado | Defecto |
| -- | --------- | ------ | --------- | ---- | ------- | ------ | ------- |
{tabla}

---

## Detalle de cada caso

"""


def generar_md(destino):
    def contar(indice, orden=None):
        d = {}
        for c in CASOS:
            clave = normaliza(c[indice])
            d[clave] = d.get(clave, 0) + 1
        if orden:
            return [(k, d[k]) for k in orden if k in d]
        return sorted(d.items(), key=lambda x: -x[1])

    estados = contar(10, ["Pasa", "Falla", "Bloqueado"])
    prios = contar(7, ["Alta", "Media", "Baja"])
    tecnicas = contar(9)

    filas_resumen = []
    for i in range(max(len(estados), len(prios), len(tecnicas))):
        def celda(lista):
            return (lista[i][0], str(lista[i][1])) if i < len(lista) else ("", "")
        e, p, t = celda(estados), celda(prios), celda(tecnicas)
        filas_resumen.append(
            f"| {e[0]} | {e[1]} | | {p[0]} | {p[1]} | | {t[0]} | {t[1]} |")

    tabla = []
    for c in CASOS:
        cid, req, titulo, _, _, _, _, prio, tipo, tec, estado, bug = c
        enlace = f"[{bug}](03-bug-reports/{bug}.md)" if bug else "—"
        tabla.append(
            f"| `{cid}` | {req} | [{titulo}](#{cid.lower()}) | {prio} | {tipo} | "
            f"{tec} | {ICONO[estado]} | {enlace} |")

    partes = [CABECERA_MD.format(
        total=len(CASOS),
        resumen="\n".join(filas_resumen),
        tabla="\n".join(tabla))]

    for c in CASOS:
        cid, req, titulo, precond, pasos, datos, esperado, prio, tipo, tec, estado, bug = c
        enlace = f"[{bug}](03-bug-reports/{bug}.md)" if bug else "—"
        pasos_md = "\n".join(f"{i}. {p}" for i, p in enumerate(pasos, 1))
        partes.append(
            f"### {cid}\n\n"
            f"**{titulo}**\n\n"
            f"| Requisito | Prioridad | Tipo | Técnica de diseño | Estado | Defecto |\n"
            f"| --------- | --------- | ---- | ----------------- | ------ | ------- |\n"
            f"| {req} | {prio} | {tipo} | {tec} | {ICONO[estado]} | {enlace} |\n\n"
            f"**Precondiciones:** {precond}\n\n"
            f"**Datos de prueba:** {datos}\n\n"
            f"**Pasos**\n\n{pasos_md}\n\n"
            f"**Resultado esperado:** {esperado}\n\n"
            f"<sub>[↑ volver al resumen](#resumen-de-la-matriz)</sub>\n"
        )

    # newline="\n" explícito: `write_text` traduciría a CRLF en Windows y el
    # fichero dejaría de ser idéntico al que genera el runner de CI en Linux.
    with open(destino, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(partes))
    print(f"  ✓ {destino.name}")


if __name__ == "__main__":
    print(f"Generando la matriz ({len(CASOS)} casos)…")
    generar_md(RAIZ / f"{BASE}.md")
    generar_csv(RAIZ / f"{BASE}.csv")
    generar_xlsx(RAIZ / f"{BASE}.xlsx")
    print("Hecho.")
